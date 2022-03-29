import datetime
import time as tm
from os import system as sys

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = Workbook()
wb = load_workbook(r'E:\devensh\project wroks\python projects\run_data.xlsx')

# grab the active worksheet
sheet = wb.active

strt = datetime.datetime.now()
#print(strt.strftime("%H:%M:%S"))
count = sheet.cell(row=2, column=1).value + 1
#print(count)
while (True):
    t = datetime.datetime.now()
    print("start time->  ", strt.strftime("%H:%M:%S"))
    elp = t - strt
    days = elp.days
    seconds = elp.seconds
    hours = seconds // 3600
    minutes = (seconds // 60) % 60
    sec = seconds % 60
    print("elapsed time->  " + str(hours) + ":" + str(minutes) + ":" + str(sec))
    print("time now->    ", t.strftime("%H:%M:%S"))
    # save in the excel file
    #print(count)
    if minutes % 10 == 0 and minutes != 0 and sec<=10:
        sheet.cell(row=count, column=2).value = strt.strftime("%Y-%m-%d , %H:%M:%S")
        sheet.cell(row=count, column=3).value = hours
        sheet.cell(row=count, column=4).value = minutes+5
        sheet.cell(row=count, column=5).value = sec
        sheet.cell(row=count, column=6).value = t.strftime("%Y-%m-%d , %H:%M:%S")
        sheet.cell(row=2, column=1).value = count
        wb.save(r"E:\devensh\project wroks\python projects\run_data.xlsx")
    tm.sleep(5)
    sys('cls')
